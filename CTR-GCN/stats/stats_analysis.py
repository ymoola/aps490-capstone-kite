from __future__ import annotations

import json
import math
import sys
import traceback
import xml.sax.saxutils as xml_utils
import zipfile
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd
import torch
import torch.nn as nn
from torch.utils.data import DataLoader

_PROJECT_ROOT = Path(__file__).resolve().parent
if str(_PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(_PROJECT_ROOT))

try:
    from .code.inference.ctr_gcn import PoseNPZDataset, build_ctr_gcn_model
except ImportError:
    from code.inference.ctr_gcn import PoseNPZDataset, build_ctr_gcn_model


PASS_FAIL_MAP = {0: "pass", 1: "fail"}
REPO_ROOT = Path(__file__).resolve().parent.parent


def _normalize_path(value: str | Path) -> str:
    return str(Path(value)).replace("\\", "/")


def _relative_to_repo(value: str | Path | None) -> str | None:
    if value in (None, ""):
        return None
    path = Path(value)
    try:
        return _normalize_path(path.resolve().relative_to(REPO_ROOT.resolve()))
    except Exception:
        return _normalize_path(path)


def _read_json(path: str | Path) -> Any:
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def _read_jsonl(path: str | Path) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    with open(path, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if line:
                rows.append(json.loads(line))
    return rows


def _compute_binary_metrics_from_predictions(y_true: np.ndarray, y_pred: np.ndarray) -> dict[str, float]:
    y_true = y_true.astype(int)
    y_pred = y_pred.astype(int)
    tp = int(((y_pred == 1) & (y_true == 1)).sum())
    tn = int(((y_pred == 0) & (y_true == 0)).sum())
    fp = int(((y_pred == 1) & (y_true == 0)).sum())
    fn = int(((y_pred == 0) & (y_true == 1)).sum())
    acc = float((y_true == y_pred).mean()) if len(y_true) else float("nan")
    tpr_fail = tp / max(1, tp + fn)
    tnr_pass = tn / max(1, tn + fp)
    return {
        "acc": acc,
        "balanced_acc": 0.5 * (tpr_fail + tnr_pass),
        "tp": tp,
        "tn": tn,
        "fp": fp,
        "fn": fn,
        "tpr_fail": tpr_fail,
        "tnr_pass": tnr_pass,
    }


def load_run_summary_table(runs_root: str | Path) -> pd.DataFrame:
    runs_root = Path(runs_root)
    records: list[dict[str, Any]] = []

    for run_dir in sorted(p for p in runs_root.iterdir() if p.is_dir() and p.name.startswith("fold")):
        summary_path = run_dir / "summary.json"
        config_path = run_dir / "run_config.json"
        if not summary_path.exists() or not config_path.exists():
            continue

        summary = _read_json(summary_path)
        config = _read_json(config_path)
        train_cfg = config["train_config"]
        history = summary.get("history", [])
        final_test = summary.get("final_test", {})
        best_val_bal_acc = max((ep["val_balanced_acc"] for ep in history), default=float("nan"))
        best_val_epoch = next(
            (ep["epoch"] for ep in history if ep["val_balanced_acc"] == best_val_bal_acc),
            None,
        )

        records.append(
            {
                "run_name": config["run_name"],
                "run_dir": str(run_dir),
                "fold": int(config["fold"]),
                "batch_size": int(train_cfg["batch_size"]),
                "lr": float(train_cfg["lr"]),
                "weight_decay": float(train_cfg["weight_decay"]),
                "epochs_requested": int(train_cfg["epochs"]),
                "device": train_cfg["device"],
                "best_ckpt": _relative_to_repo(summary["paths"].get("best_ckpt")),
                "train_npz": _relative_to_repo(summary["paths"].get("train_npz")),
                "val_npz": _relative_to_repo(summary["paths"].get("val_npz")),
                "test_npz": _relative_to_repo(summary["paths"].get("test_npz")),
                "config_path": _relative_to_repo(config_path),
                "summary_path": _relative_to_repo(summary_path),
                "hparam_key": f"bs{int(train_cfg['batch_size'])}_lr{float(train_cfg['lr']):.0e}_wd{float(train_cfg['weight_decay']):.0e}",
                "best_val_bal_acc": best_val_bal_acc,
                "best_val_epoch": best_val_epoch,
                "total_epochs": len(history),
                "test_bal_acc": final_test.get("balanced_acc", float("nan")),
                "test_acc": final_test.get("acc", float("nan")),
                "test_tp": final_test.get("tp", float("nan")),
                "test_tn": final_test.get("tn", float("nan")),
                "test_fp": final_test.get("fp", float("nan")),
                "test_fn": final_test.get("fn", float("nan")),
                "test_tpr_fail": final_test.get("tpr_fail", float("nan")),
                "test_tnr_pass": final_test.get("tnr_pass", float("nan")),
                "test_loss": final_test.get("loss", float("nan")),
            }
        )

    if not records:
        raise FileNotFoundError(f"No run summaries found under {runs_root}")

    return pd.DataFrame(records).sort_values(["fold", "best_val_bal_acc", "run_name"], ascending=[True, False, True])


def select_best_per_fold(runs_df: pd.DataFrame) -> pd.DataFrame:
    best = (
        runs_df.sort_values(["fold", "best_val_bal_acc", "run_name"], ascending=[True, False, True])
        .groupby("fold", as_index=False, sort=True)
        .first()
    )
    return best.sort_values("fold").reset_index(drop=True)


def load_test_split_metadata(cv_splits_dir: str | Path, folds: list[int] | None = None) -> pd.DataFrame:
    cv_splits_dir = Path(cv_splits_dir)
    folds = list(range(5)) if folds is None else folds
    rows: list[dict[str, Any]] = []

    for fold in folds:
        for row in _read_jsonl(cv_splits_dir / f"fold_{fold}_test.jsonl"):
            row = dict(row)
            row["fold"] = fold
            row["npz_path_norm"] = _normalize_path(row["npz_path"])
            rows.append(row)

    df = pd.DataFrame(rows)
    if df.empty:
        raise FileNotFoundError(f"No fold test metadata found under {cv_splits_dir}")
    return df


def _make_eval_loader(ds: PoseNPZDataset, batch_size: int, num_workers: int) -> DataLoader:
    use_pin_memory = torch.cuda.is_available()
    return DataLoader(
        ds,
        batch_size=batch_size,
        shuffle=False,
        num_workers=0 if num_workers else 0,
        pin_memory=use_pin_memory,
        drop_last=False,
    )


@torch.no_grad()
def evaluate_run_per_sample(run_row: pd.Series, device: str | None = None) -> pd.DataFrame:
    config = _read_json(REPO_ROOT / run_row["config_path"])
    model_kwargs = config["model_kwargs"]
    train_cfg = config["train_config"]
    ckpt_path = REPO_ROOT / run_row["best_ckpt"]
    test_npz = REPO_ROOT / run_row["test_npz"]

    if not ckpt_path.exists():
        raise FileNotFoundError(f"Checkpoint not found: {ckpt_path}")

    ds = PoseNPZDataset(test_npz)
    batch_size = int(train_cfg.get("batch_size", 32))
    num_workers = int(train_cfg.get("num_workers", 0))
    chosen_device = device or ("cuda" if torch.cuda.is_available() else "cpu")
    torch_device = torch.device(chosen_device)

    model = build_ctr_gcn_model(
        ctr_repo_root=_PROJECT_ROOT / "frameworks" / "CTR-GCN",
        **model_kwargs,
    ).to(torch_device)
    checkpoint = torch.load(ckpt_path, map_location=torch_device)
    model.load_state_dict(checkpoint["model_state"])
    model.eval()

    criterion = nn.CrossEntropyLoss()
    loader = _make_eval_loader(ds, batch_size=batch_size, num_workers=num_workers)

    rows: list[dict[str, Any]] = []
    start_idx = 0
    total_loss = 0.0

    for x, y in loader:
        x = x.to(torch_device, non_blocking=True)
        y = y.to(torch_device, non_blocking=True)
        logits = model(x)
        probs = torch.softmax(logits, dim=1)
        loss = criterion(logits, y)
        preds = torch.argmax(logits, dim=1)

        batch_size_now = x.size(0)
        total_loss += float(loss.item()) * batch_size_now

        logits_cpu = logits.detach().cpu().numpy()
        probs_cpu = probs.detach().cpu().numpy()
        y_cpu = y.detach().cpu().numpy()
        preds_cpu = preds.detach().cpu().numpy()

        for offset in range(batch_size_now):
            sample_idx = start_idx + offset
            meta = ds.get_meta(sample_idx) or {}
            rows.append(
                {
                    "fold": int(run_row["fold"]),
                    "run_name": run_row["run_name"],
                    "hparam_key": run_row["hparam_key"],
                    "batch_size": int(run_row["batch_size"]),
                    "lr": float(run_row["lr"]),
                    "weight_decay": float(run_row["weight_decay"]),
                    "best_val_bal_acc": float(run_row["best_val_bal_acc"]),
                    "best_val_epoch": int(run_row["best_val_epoch"]),
                    "total_epochs": int(run_row["total_epochs"]),
                    "best_ckpt": run_row["best_ckpt"],
                    "test_npz": run_row["test_npz"],
                    "sample_index_within_fold_test": sample_idx,
                    "npz_path": meta.get("npz_path"),
                    "npz_path_rel": _relative_to_repo(meta.get("npz_path")),
                    "npz_path_norm": _normalize_path(meta.get("npz_path", "")),
                    "participant_dir": meta.get("participant_dir"),
                    "participant_key": meta.get("participant_key"),
                    "angle_from_npz": meta.get("angle"),
                    "label_code_from_npz": meta.get("label_code"),
                    "true_class": int(y_cpu[offset]),
                    "pred_class": int(preds_cpu[offset]),
                    "true_binary": PASS_FAIL_MAP[int(y_cpu[offset])],
                    "pred_binary": PASS_FAIL_MAP[int(preds_cpu[offset])],
                    "correct": bool(y_cpu[offset] == preds_cpu[offset]),
                    "prob_pass": float(probs_cpu[offset, 0]),
                    "prob_fail": float(probs_cpu[offset, 1]),
                    "logit_pass": float(logits_cpu[offset, 0]),
                    "logit_fail": float(logits_cpu[offset, 1]),
                }
            )

        start_idx += batch_size_now

    sample_df = pd.DataFrame(rows)
    metrics = _compute_binary_metrics_from_predictions(
        sample_df["true_class"].to_numpy(),
        sample_df["pred_class"].to_numpy(),
    )
    sample_df.attrs["metrics"] = {
        **metrics,
        "loss": total_loss / max(1, len(ds)),
    }
    return sample_df


def build_master_performance_table(
    runs_root: str | Path = str(_PROJECT_ROOT / "runs" / "ctr_gcn_kfold_hpo"),
    cv_splits_dir: str | Path = str(_PROJECT_ROOT / "data" / "cv_splits"),
    device: str | None = None,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    runs_df = load_run_summary_table(runs_root)
    best_per_fold = select_best_per_fold(runs_df)
    split_df = load_test_split_metadata(cv_splits_dir, folds=best_per_fold["fold"].tolist())

    per_fold_tables = []
    for _, row in best_per_fold.iterrows():
        pred_df = evaluate_run_per_sample(row, device=device)
        merged = pred_df.merge(
            split_df[split_df["fold"] == int(row["fold"])],
            on=["fold", "npz_path_norm"],
            how="left",
            suffixes=("", "_split"),
        )
        per_fold_tables.append(merged)

    master_df = pd.concat(per_fold_tables, ignore_index=True)
    master_df["error"] = ~master_df["correct"]

    if len(master_df) != 1623:
        raise ValueError(f"Expected 1623 total test rows, found {len(master_df)}")

    ordered_cols = [
        "fold",
        "run_name",
        "hparam_key",
        "batch_size",
        "lr",
        "weight_decay",
        "best_val_bal_acc",
        "best_val_epoch",
        "total_epochs",
        "best_ckpt",
        "test_npz",
        "sample_index_within_fold_test",
        "rel_path",
        "participant_dir",
        "participant_id",
        "participant_key",
        "footwear_id",
        "label_code",
        "label_binary",
        "slope_dir",
        "angle_deg",
        "time_str",
        "stage_tag",
        "true_class",
        "pred_class",
        "true_binary",
        "pred_binary",
        "correct",
        "error",
        "prob_pass",
        "prob_fail",
        "logit_pass",
        "logit_fail",
        "angle_from_npz",
        "label_code_from_npz",
    ]
    existing_cols = [col for col in ordered_cols if col in master_df.columns]
    trailing_cols = [col for col in master_df.columns if col not in existing_cols]
    master_df = master_df[existing_cols + trailing_cols]
    master_df = master_df.drop(
        columns=[
            col
            for col in ["npz_path", "npz_path_norm", "npz_path_rel", "npz_path_split"]
            if col in master_df.columns
        ]
    )

    return master_df, best_per_fold


def summarize_group_success(master_df: pd.DataFrame, group_cols: str | list[str]) -> pd.DataFrame:
    if isinstance(group_cols, str):
        group_cols = [group_cols]

    summary = (
        master_df.groupby(group_cols, dropna=False)
        .agg(
            videos=("correct", "size"),
            correct_count=("correct", "sum"),
            error_count=("error", "sum"),
            success_rate=("correct", "mean"),
            error_rate=("error", "mean"),
            mean_prob_fail=("prob_fail", "mean"),
        )
        .reset_index()
        .sort_values(group_cols)
    )
    return summary


def make_standard_summary_tables(master_df: pd.DataFrame) -> dict[str, pd.DataFrame]:
    return {
        "by_participant": summarize_group_success(master_df, "participant_id"),
        "by_footwear": summarize_group_success(master_df, "footwear_id"),
        "by_direction": summarize_group_success(master_df, "slope_dir"),
        "by_angle": summarize_group_success(master_df, "angle_deg"),
        "by_label_code": summarize_group_success(master_df, "label_code"),
        "by_footwear_direction": summarize_group_success(master_df, ["footwear_id", "slope_dir"]),
        "by_participant_direction": summarize_group_success(master_df, ["participant_id", "slope_dir"]),
    }


def export_analysis_excel(
    master_df: pd.DataFrame,
    best_per_fold: pd.DataFrame,
    summary_tables: dict[str, pd.DataFrame],
    output_path: str | Path,
) -> Path:
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    sheets: list[tuple[str, pd.DataFrame]] = [
        ("per_video", master_df),
        ("best_per_fold", best_per_fold),
        *[(sheet_name[:31], table) for sheet_name, table in summary_tables.items()],
    ]

    engine = None
    try:
        import openpyxl  # noqa: F401

        engine = "openpyxl"
    except Exception:
        try:
            import xlsxwriter  # noqa: F401

            engine = "xlsxwriter"
        except Exception:
            engine = None

    if engine is not None:
        with pd.ExcelWriter(output_path, engine=engine) as writer:
            for sheet_name, table in sheets:
                table.to_excel(writer, sheet_name=sheet_name, index=False)

            if engine == "openpyxl":
                _add_native_excel_charts(writer.book, sheets)
    else:
        _write_simple_xlsx(output_path, sheets)

    return output_path


def _excel_col_name(col_idx_zero_based: int) -> str:
    col_idx = col_idx_zero_based + 1
    letters = []
    while col_idx:
        col_idx, rem = divmod(col_idx - 1, 26)
        letters.append(chr(65 + rem))
    return "".join(reversed(letters))


def _xml_cell(cell_ref: str, value: Any) -> str:
    if pd.isna(value):
        return f'<c r="{cell_ref}"/>'
    if isinstance(value, (bool, np.bool_)):
        return f'<c r="{cell_ref}" t="b"><v>{1 if value else 0}</v></c>'
    if isinstance(value, (int, float, np.integer, np.floating)) and not isinstance(value, bool):
        if isinstance(value, float) and (math.isnan(value) or math.isinf(value)):
            return f'<c r="{cell_ref}"/>'
        return f'<c r="{cell_ref}"><v>{value}</v></c>'

    escaped = xml_utils.escape(str(value))
    return f'<c r="{cell_ref}" t="inlineStr"><is><t>{escaped}</t></is></c>'


def _worksheet_xml(df: pd.DataFrame) -> str:
    rows_xml: list[str] = []
    header_cells = []
    for col_idx, col_name in enumerate(df.columns):
        cell_ref = f"{_excel_col_name(col_idx)}1"
        header_cells.append(_xml_cell(cell_ref, col_name))
    rows_xml.append(f'<row r="1">{"".join(header_cells)}</row>')

    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        cell_xml = []
        for col_idx, value in enumerate(row):
            cell_ref = f"{_excel_col_name(col_idx)}{row_idx}"
            cell_xml.append(_xml_cell(cell_ref, value))
        rows_xml.append(f'<row r="{row_idx}">{"".join(cell_xml)}</row>')

    dimension_ref = f"A1:{_excel_col_name(max(len(df.columns) - 1, 0))}{max(len(df) + 1, 1)}"
    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
        f'<dimension ref="{dimension_ref}"/>'
        '<sheetViews><sheetView workbookViewId="0"/></sheetViews>'
        '<sheetFormatPr defaultRowHeight="15"/>'
        f'<sheetData>{"".join(rows_xml)}</sheetData>'
        '</worksheet>'
    )


def _write_simple_xlsx(output_path: Path, sheets: list[tuple[str, pd.DataFrame]]) -> None:
    workbook_sheets = []
    workbook_rels = []
    content_type_overrides = []

    for idx, (sheet_name, _) in enumerate(sheets, start=1):
        escaped_name = xml_utils.escape(sheet_name)
        workbook_sheets.append(
            f'<sheet name="{escaped_name}" sheetId="{idx}" r:id="rId{idx}"/>'
        )
        workbook_rels.append(
            f'<Relationship Id="rId{idx}" '
            'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" '
            f'Target="worksheets/sheet{idx}.xml"/>'
        )
        content_type_overrides.append(
            f'<Override PartName="/xl/worksheets/sheet{idx}.xml" '
            'ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>'
        )

    workbook_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
        'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
        '<bookViews><workbookView/></bookViews>'
        f'<sheets>{"".join(workbook_sheets)}</sheets>'
        '</workbook>'
    )
    workbook_rels_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        f'{"".join(workbook_rels)}'
        '</Relationships>'
    )
    root_rels_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId1" '
        'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" '
        'Target="xl/workbook.xml"/>'
        '<Relationship Id="rId2" '
        'Type="http://schemas.openxmlformats.org/package/2006/relationships/metadata/core-properties" '
        'Target="docProps/core.xml"/>'
        '<Relationship Id="rId3" '
        'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/extended-properties" '
        'Target="docProps/app.xml"/>'
        '</Relationships>'
    )
    content_types_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
        '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
        '<Default Extension="xml" ContentType="application/xml"/>'
        '<Override PartName="/xl/workbook.xml" '
        'ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>'
        '<Override PartName="/docProps/core.xml" '
        'ContentType="application/vnd.openxmlformats-package.core-properties+xml"/>'
        '<Override PartName="/docProps/app.xml" '
        'ContentType="application/vnd.openxmlformats-officedocument.extended-properties+xml"/>'
        f'{"".join(content_type_overrides)}'
        '</Types>'
    )
    app_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Properties xmlns="http://schemas.openxmlformats.org/officeDocument/2006/extended-properties" '
        'xmlns:vt="http://schemas.openxmlformats.org/officeDocument/2006/docPropsVTypes">'
        '<Application>Codex</Application>'
        '</Properties>'
    )
    core_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<cp:coreProperties xmlns:cp="http://schemas.openxmlformats.org/package/2006/metadata/core-properties" '
        'xmlns:dc="http://purl.org/dc/elements/1.1/" '
        'xmlns:dcterms="http://purl.org/dc/terms/" '
        'xmlns:dcmitype="http://purl.org/dc/dcmitype/" '
        'xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance">'
        '<dc:creator>Codex</dc:creator>'
        '<cp:lastModifiedBy>Codex</cp:lastModifiedBy>'
        '<dc:title>Performance Stats Analysis</dc:title>'
        '</cp:coreProperties>'
    )

    with zipfile.ZipFile(output_path, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("[Content_Types].xml", content_types_xml)
        zf.writestr("_rels/.rels", root_rels_xml)
        zf.writestr("docProps/app.xml", app_xml)
        zf.writestr("docProps/core.xml", core_xml)
        zf.writestr("xl/workbook.xml", workbook_xml)
        zf.writestr("xl/_rels/workbook.xml.rels", workbook_rels_xml)
        for idx, (_, df) in enumerate(sheets, start=1):
            zf.writestr(f"xl/worksheets/sheet{idx}.xml", _worksheet_xml(df))


def _add_native_excel_charts(workbook: Any, sheets: list[tuple[str, pd.DataFrame]]) -> None:
    from openpyxl.chart import BarChart, LineChart, Reference
    from openpyxl.chart.label import DataLabelList

    sheet_tables = {name: df for name, df in sheets}

    for sheet_name, df in sheet_tables.items():
        if sheet_name == "per_video" or df.empty:
            continue

        ws = workbook[sheet_name]
        max_row = len(df) + 1
        max_col = len(df.columns)

        if sheet_name == "best_per_fold":
            fold_col = df.columns.get_loc("fold") + 1
            bal_acc_col = df.columns.get_loc("test_bal_acc") + 1
            acc_col = df.columns.get_loc("test_acc") + 1
            _add_chart_to_sheet(
                ws,
                chart_type="col",
                title="Test Balanced Accuracy by Fold",
                categories_col=fold_col,
                data_cols=[bal_acc_col],
                anchor="L2",
                max_row=max_row,
                y_title="Balanced Accuracy",
            )
            _add_chart_to_sheet(
                ws,
                chart_type="col",
                title="Test Accuracy by Fold",
                categories_col=fold_col,
                data_cols=[acc_col],
                anchor="L20",
                max_row=max_row,
                y_title="Accuracy",
            )
            continue

        if "success_rate" in df.columns:
            category_col = 1
            if sheet_name in {"by_footwear_direction", "by_participant_direction"}:
                _write_combo_category_labels(ws, df, [1, 2], "combo_label", max_col + 1)
                category_col = max_col + 1

            chart_type = "line" if sheet_name == "by_angle" else "col"
            _add_chart_to_sheet(
                ws,
                chart_type=chart_type,
                title=f"{sheet_name} Success Rate",
                categories_col=category_col,
                data_cols=[df.columns.get_loc("success_rate") + 1],
                anchor="J2",
                max_row=max_row,
                y_title="Success Rate",
            )

            if "videos" in df.columns:
                _add_chart_to_sheet(
                    ws,
                    chart_type="col",
                    title=f"{sheet_name} Video Count",
                    categories_col=category_col,
                    data_cols=[df.columns.get_loc("videos") + 1],
                    anchor="J20",
                    max_row=max_row,
                    y_title="Videos",
                )


def _write_combo_category_labels(
    ws: Any,
    df: pd.DataFrame,
    category_cols: list[int],
    header: str,
    output_col: int,
) -> None:
    ws.cell(row=1, column=output_col, value=header)
    for row_idx in range(2, len(df) + 2):
        parts = [str(ws.cell(row=row_idx, column=col).value) for col in category_cols]
        ws.cell(row=row_idx, column=output_col, value=" | ".join(parts))


def _add_chart_to_sheet(
    ws: Any,
    chart_type: str,
    title: str,
    categories_col: int,
    data_cols: list[int],
    anchor: str,
    max_row: int,
    y_title: str,
) -> None:
    from openpyxl.chart import BarChart, LineChart, Reference
    from openpyxl.chart.label import DataLabelList

    if chart_type == "line":
        chart = LineChart()
    else:
        chart = BarChart()
        chart.type = "col"
        chart.style = 10

    chart.title = title
    chart.y_axis.title = y_title
    chart.x_axis.title = "Category"
    chart.height = 8
    chart.width = 14

    cats = Reference(ws, min_col=categories_col, min_row=2, max_row=max_row)
    for data_col in data_cols:
        data = Reference(ws, min_col=data_col, min_row=1, max_row=max_row)
        chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)

    if chart_type != "line":
        chart.dLbls = DataLabelList()
        chart.dLbls.showVal = True

    ws.add_chart(chart, anchor)



def run_full_analysis(
    runs_root: str | Path = str(_PROJECT_ROOT / "runs" / "ctr_gcn_kfold_hpo"),
    cv_splits_dir: str | Path = str(_PROJECT_ROOT / "data" / "cv_splits"),
    analysis_dir: str | Path = str(_PROJECT_ROOT / "analysis_outputs"),
    device: str | None = None,
) -> dict[str, Any]:
    analysis_dir = Path(analysis_dir)
    analysis_dir.mkdir(parents=True, exist_ok=True)

    master_df, best_per_fold = build_master_performance_table(
        runs_root=runs_root,
        cv_splits_dir=cv_splits_dir,
        device=device,
    )
    summary_tables = make_standard_summary_tables(master_df)
    excel_path = export_analysis_excel(
        master_df=master_df,
        best_per_fold=best_per_fold,
        summary_tables=summary_tables,
        output_path=analysis_dir / "performance_stats_analysis.xlsx",
    )

    return {
        "master_df": master_df,
        "best_per_fold": best_per_fold,
        "summary_tables": summary_tables,
        "excel_path": excel_path,
    }


if __name__ == "__main__":
    try:
        results = run_full_analysis()
        print(f"Wrote workbook: {results['excel_path']}")
        print(f"Per-video rows: {len(results['master_df'])}")
        print("Best model per fold:")
        print(results["best_per_fold"][["fold", "hparam_key", "test_bal_acc", "test_acc"]].to_string(index=False))
    except Exception as exc:
        print("Stats analysis failed.")
        print(f"{type(exc).__name__}: {exc}")
        traceback.print_exc()
        raise
