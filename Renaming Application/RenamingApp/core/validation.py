"""
Integrated CV model validation for the SlopeSense renaming application.

After videos are renamed, this module classifies each video as Pass/Fail
using the CTR-GCN skeleton-based action recognition model.

Pipeline per video:
  MP4 -> YOLO pose extraction -> interpolation -> smoothing ->
  temporal resampling (T=100) -> normalization -> CTR-GCN inference -> Pass/Fail

Runtime backends:
  - Cross-platform default: YOLO ONNX + CTR-GCN ONNX Runtime
  - macOS optimized: direct CoreML YOLO + CTR-GCN ONNX Runtime

No PyTorch is required at runtime.
"""

from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path
import platform
from typing import Callable, List, Optional, Tuple

import cv2
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# ===================================================================
# Configuration
# ===================================================================
@dataclass
class ValidationConfig:
    """All knobs needed to run the validation pipeline."""

    yolo_model_path: str          # path to yolo26x-pose.onnx/.mlpackage/.mlmodel
    ctr_gcn_checkpoint_path: str  # path to classifier.onnx

    # Execution backend preference:
    #   auto -> CUDA, then CPU
    #   coreml -> explicit opt-in only (kept for debugging; not used by default)
    device: str = "auto"

    # Pose extraction
    yolo_batch_size: int = 8      # kept for interface compatibility; ONNX runs per-frame
    conf_thr: float = 0.05

    # Preprocessing
    do_interp: bool = True
    do_smooth: bool = True
    fps_scale: int = 4
    ema_alpha: float = 0.7

    # CTR-GCN input shape
    fixed_t: int = 100
    num_class: int = 2
    num_point: int = 17
    num_person: int = 1
    in_channels: int = 3


@dataclass
class ValidationResult:
    """Result of validating a single renamed video."""

    original_video: str
    renamed_video: str
    renamed_video_path: str
    tipper_label: str       # from tipper filename: Pass / Fail / Undecided
    predicted_label: str    # from model: Pass / Fail
    predicted_prob: float   # softmax confidence
    labels_match: bool
    error: Optional[str] = None


def _is_coreml_model_path(model_path: Optional[str]) -> bool:
    if not model_path:
        return False
    return Path(model_path).suffix.lower() in {".mlpackage", ".mlmodel"}


def _has_coremltools() -> bool:
    if platform.system() != "Darwin":
        return False
    try:
        import coremltools  # noqa: F401
    except Exception:
        return False
    return True


def resolve_onnx_execution(
    requested_device: str = "auto",
) -> Tuple[str, List[str]]:
    """
    Resolve the effective ONNX Runtime backend and provider chain.

    Fallback order:
      auto -> CUDA -> CPU
      coreml -> explicit opt-in only
    """
    import onnxruntime as ort

    available = set(ort.get_available_providers())
    requested = requested_device.lower().strip()

    if requested in ("auto", "cuda") and "CUDAExecutionProvider" in available:
        return "cuda", ["CUDAExecutionProvider", "CPUExecutionProvider"]
    if requested in ("coreml", "mac", "mps") and "CoreMLExecutionProvider" in available:
        return "coreml", ["CoreMLExecutionProvider", "CPUExecutionProvider"]
    return "cpu", ["CPUExecutionProvider"]


def resolve_validation_execution(
    requested_device: str = "auto",
    yolo_model_path: Optional[str] = None,
) -> Tuple[Tuple[str, List[str]], Tuple[str, List[str]]]:
    """
    Resolve per-model execution backends for validation.

    Policy:
      auto on CUDA-capable systems -> YOLO cuda, CTR-GCN cuda
      auto on macOS with CoreML YOLO model -> YOLO direct-coreml, CTR-GCN coreml (if available)
      auto on macOS without CoreML YOLO model -> YOLO cpu, CTR-GCN coreml (if available)
      auto otherwise -> YOLO cpu, CTR-GCN cpu

    Explicit requests still force both models to the same backend family.
    """
    import onnxruntime as ort

    available = set(ort.get_available_providers())
    requested = requested_device.lower().strip()
    has_direct_coreml_yolo = _has_coremltools() and _is_coreml_model_path(yolo_model_path)

    if requested in ("auto", "cuda") and "CUDAExecutionProvider" in available:
        cuda = ("cuda", ["CUDAExecutionProvider", "CPUExecutionProvider"])
        return cuda, cuda

    if requested == "auto" and platform.system() == "Darwin":
        yolo = ("coreml_direct", ["CoreMLDirect"]) if has_direct_coreml_yolo else ("cpu", ["CPUExecutionProvider"])
        if "CoreMLExecutionProvider" in available:
            ctr = ("coreml", ["CoreMLExecutionProvider", "CPUExecutionProvider"])
        else:
            ctr = ("cpu", ["CPUExecutionProvider"])
        return yolo, ctr

    if requested in ("coreml", "mac", "mps") and "CoreMLExecutionProvider" in available:
        if has_direct_coreml_yolo:
            yolo = ("coreml_direct", ["CoreMLDirect"])
        else:
            yolo = ("coreml", ["CoreMLExecutionProvider", "CPUExecutionProvider"])
        coreml = ("coreml", ["CoreMLExecutionProvider", "CPUExecutionProvider"])
        return yolo, coreml

    cpu = ("cpu", ["CPUExecutionProvider"])
    return cpu, cpu


def _onnx_provider_candidates(
    requested_device: str = "auto",
) -> List[Tuple[str, List[str]]]:
    import onnxruntime as ort

    available = set(ort.get_available_providers())
    requested = requested_device.lower().strip()
    candidates: List[Tuple[str, List[str]]] = []

    if requested in ("auto", "cuda") and "CUDAExecutionProvider" in available:
        candidates.append(("cuda", ["CUDAExecutionProvider", "CPUExecutionProvider"]))
    if requested in ("coreml", "mac", "mps") and "CoreMLExecutionProvider" in available:
        candidates.append(("coreml", ["CoreMLExecutionProvider", "CPUExecutionProvider"]))
    if not candidates or requested == "cpu":
        candidates.append(("cpu", ["CPUExecutionProvider"]))
    elif candidates[-1][0] != "cpu":
        candidates.append(("cpu", ["CPUExecutionProvider"]))
    return candidates


def create_onnx_session(
    model_path: str,
    requested_device: str = "auto",
    log: Optional[Callable[[str], None]] = None,
) -> Tuple[object, str, List[str]]:
    """Create an ONNX session with provider fallback."""
    import onnxruntime as ort

    errors: List[str] = []
    for backend, providers in _onnx_provider_candidates(requested_device):
        try:
            session = ort.InferenceSession(model_path, providers=providers)
            return session, backend, providers
        except Exception as exc:
            errors.append(f"{backend}: {exc}")
            if log is not None:
                log(
                    f"[Validate] Provider init failed for {Path(model_path).name} "
                    f"on {backend}; trying next fallback."
                )

    raise RuntimeError(
        f"Could not initialize ONNX session for {Path(model_path).name}. "
        f"Tried: {' | '.join(errors)}"
    )


# ===================================================================
# YOLO ONNX inference helpers
# ===================================================================
def _letterbox(
    img: np.ndarray, imgsz: int = 640
) -> Tuple[np.ndarray, float, int, int]:
    """Resize + pad image to imgsz x imgsz. Returns (padded_img, ratio, pad_top, pad_left)."""
    h, w = img.shape[:2]
    r = min(imgsz / h, imgsz / w)
    new_w, new_h = int(round(w * r)), int(round(h * r))
    img = cv2.resize(img, (new_w, new_h), interpolation=cv2.INTER_LINEAR)
    pad_h = imgsz - new_h
    pad_w = imgsz - new_w
    top = pad_h // 2
    left = pad_w // 2
    img = cv2.copyMakeBorder(
        img, top, pad_h - top, left, pad_w - left,
        cv2.BORDER_CONSTANT, value=(114, 114, 114),
    )
    return img, r, top, left


def _nms(boxes: np.ndarray, scores: np.ndarray, iou_thr: float) -> List[int]:
    """CPU non-maximum suppression. boxes: (N, 4) xyxy."""
    order = scores.argsort()[::-1]
    keep: List[int] = []
    while order.size > 0:
        i = int(order[0])
        keep.append(i)
        if order.size == 1:
            break
        xx1 = np.maximum(boxes[i, 0], boxes[order[1:], 0])
        yy1 = np.maximum(boxes[i, 1], boxes[order[1:], 1])
        xx2 = np.minimum(boxes[i, 2], boxes[order[1:], 2])
        yy2 = np.minimum(boxes[i, 3], boxes[order[1:], 3])
        inter = np.maximum(0.0, xx2 - xx1) * np.maximum(0.0, yy2 - yy1)
        area_i = (boxes[i, 2] - boxes[i, 0]) * (boxes[i, 3] - boxes[i, 1])
        area_j = (
            (boxes[order[1:], 2] - boxes[order[1:], 0])
            * (boxes[order[1:], 3] - boxes[order[1:], 1])
        )
        iou = inter / (area_i + area_j - inter + 1e-6)
        order = order[1:][iou < iou_thr]
    return keep


def _yolo_postprocess(
    raw: np.ndarray,
    ratio: float,
    pad_top: int,
    pad_left: int,
    conf_thr: float = 0.05,
    iou_thr: float = 0.45,
) -> List[np.ndarray]:
    """
    Decode raw YOLOv8-pose ONNX output into a list of per-person keypoint arrays.

    Parameters
    ----------
    raw
        Direct output of the ONNX session. Supports two layouts:
          1. Legacy anchor-style export: (1, 56, N) or (1, N, 56)
          2. End-to-end NMS export:      (1, N, 6 + 17*3)
    ratio, pad_top, pad_left
        Letterbox parameters from _letterbox().

    Returns
    -------
    List of (17, 3) float32 arrays [x_orig, y_orig, conf] in original image coords.
    """
    # Newer Ultralytics ONNX exports can include post-NMS detections directly:
    # [x1, y1, x2, y2, score, class_id, kpts...]. In this case coordinates remain
    # in letterboxed image space and only need to be scaled back to the original frame.
    if raw.ndim == 3 and raw.shape[0] == 1 and raw.shape[2] >= 6:
        det_dim = raw.shape[2]
        kpt_values = det_dim - 6
        if kpt_values > 0 and kpt_values % 3 == 0:
            num_kpts = kpt_values // 3
            pred = raw[0]
            pred = pred[pred[:, 4] > conf_thr]
            detections: List[np.ndarray] = []
            for det in pred:
                kpts = det[6:].reshape(num_kpts, 3).copy()
                kpts[:, 0] = (kpts[:, 0] - pad_left) / ratio
                kpts[:, 1] = (kpts[:, 1] - pad_top) / ratio
                detections.append(kpts.astype(np.float32))
            return detections

    # Support both (1, 56, N) and (1, N, 56)
    if raw.ndim == 3 and raw.shape[1] == 56:
        pred = raw[0].T          # (N, 56)
    elif raw.ndim == 3 and raw.shape[2] == 56:
        pred = raw[0]            # (N, 56)
    else:
        return []

    scores = pred[:, 4]
    mask = scores > conf_thr
    pred = pred[mask]
    if len(pred) == 0:
        return []

    cx, cy, w, h = pred[:, 0], pred[:, 1], pred[:, 2], pred[:, 3]
    boxes = np.stack([cx - w / 2, cy - h / 2, cx + w / 2, cy + h / 2], axis=1)
    keep = _nms(boxes, pred[:, 4], iou_thr)
    pred = pred[keep]

    detections: List[np.ndarray] = []
    for det in pred:
        kpts = det[5:].reshape(17, 3).copy()  # x, y, vis in letterbox-640 space
        kpts[:, 0] = (kpts[:, 0] - pad_left) / ratio
        kpts[:, 1] = (kpts[:, 1] - pad_top) / ratio
        detections.append(kpts.astype(np.float32))
    return detections


class YoloPoseONNX:
    """ONNX-Runtime wrapper for a YOLOv8-pose model."""

    def __init__(
        self,
        model_path: str,
        imgsz: int = 640,
        providers: Optional[List[str]] = None,
        session=None,
    ) -> None:
        import onnxruntime as ort
        if session is not None:
            self._session = session
        else:
            if providers is None:
                providers = ["CPUExecutionProvider"]
            self._session = ort.InferenceSession(model_path, providers=providers)
        self._input_name = self._session.get_inputs()[0].name
        self._imgsz = imgsz

    def predict_frame(
        self, frame: np.ndarray, conf_thr: float = 0.05
    ) -> List[np.ndarray]:
        """Run inference on one BGR frame. Returns list of (17, 3) keypoint arrays."""
        img, ratio, pad_top, pad_left = _letterbox(frame, self._imgsz)
        inp = (
            cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
            .transpose(2, 0, 1)[np.newaxis]
            .astype(np.float32)
            / 255.0
        )
        raw = self._session.run(None, {self._input_name: inp})[0]
        return _yolo_postprocess(raw, ratio, pad_top, pad_left, conf_thr)


class YoloPoseCoreML:
    """Direct CoreML wrapper for a YOLO pose .mlpackage/.mlmodel export."""

    def __init__(self, model_path: str, imgsz: int = 640) -> None:
        import coremltools as ct

        self._model = ct.models.MLModel(model_path)
        self._input_name = self._model.get_spec().description.input[0].name
        self._imgsz = imgsz

    def predict_frame(
        self, frame: np.ndarray, conf_thr: float = 0.05
    ) -> List[np.ndarray]:
        from PIL import Image

        img, ratio, pad_top, pad_left = _letterbox(frame, self._imgsz)
        pil = Image.fromarray(cv2.cvtColor(img, cv2.COLOR_BGR2RGB))
        output = self._model.predict({self._input_name: pil})
        raw = np.asarray(next(iter(output.values())), dtype=np.float32)
        return _yolo_postprocess(raw, ratio, pad_top, pad_left, conf_thr)


# ===================================================================
# Person selection helpers
# ===================================================================
def _skeleton_center(
    xy: np.ndarray, conf: Optional[np.ndarray], conf_thr: float,
) -> np.ndarray:
    if conf is None:
        return np.nanmean(xy, axis=0)
    m = conf >= conf_thr
    if m.sum() < 2:
        return np.nanmean(xy, axis=0)
    return xy[m].mean(axis=0)


def _skeleton_area(
    xy: np.ndarray, conf: Optional[np.ndarray], conf_thr: float,
) -> float:
    pts = xy if conf is None else xy[conf >= conf_thr]
    if len(pts) < 2:
        return 0.0
    x1, y1 = np.min(pts[:, 0]), np.min(pts[:, 1])
    x2, y2 = np.max(pts[:, 0]), np.max(pts[:, 1])
    if not np.isfinite([x1, y1, x2, y2]).all():
        return 0.0
    return float(max(0.0, x2 - x1) * max(0.0, y2 - y1))


def _pick_single_person(
    detections: List[np.ndarray],
    num_kpts: int = 17,
    prev_center: Optional[np.ndarray] = None,
    conf_thr: float = 0.05,
    width: int = 1920,
    height: int = 1080,
) -> Tuple[np.ndarray, Optional[np.ndarray]]:
    """
    Select the most likely single person from ONNX-decoded detections.

    Parameters
    ----------
    detections : list of (V, 3) float32 arrays [x, y, conf] in original image coords.
    """
    if not detections:
        return np.zeros((num_kpts, 3), dtype=np.float32), prev_center

    if len(detections) == 1:
        kpts = detections[0]
        center = _skeleton_center(kpts[:, :2], kpts[:, 2], conf_thr)
        return kpts, center

    M = len(detections)
    xy = np.stack([d[:, :2] for d in detections])   # (M, V, 2)
    conf = np.stack([d[:, 2] for d in detections])  # (M, V)

    centers = np.stack(
        [_skeleton_center(xy[i], conf[i], conf_thr) for i in range(M)]
    )
    areas = np.array(
        [_skeleton_area(xy[i], conf[i], conf_thr) for i in range(M)],
        dtype=np.float32,
    )
    mean_conf = conf.mean(axis=1).astype(np.float32)

    size_norm = areas / (areas.max() + 1e-6) if areas.max() > 0 else np.zeros(M)
    mc_range = mean_conf.max() - mean_conf.min()
    conf_norm = (
        (mean_conf - mean_conf.min()) / (mc_range + 1e-6)
        if mc_range > 0
        else np.zeros(M)
    )

    if prev_center is not None and np.isfinite(prev_center).all():
        diag = float(np.sqrt(width ** 2 + height ** 2))
        dist = np.linalg.norm(centers - prev_center[None, :], axis=1)
        track_score = 1.0 - np.clip(dist / (diag + 1e-6), 0, 1)
    else:
        track_score = np.zeros(M, dtype=np.float32)

    score = 0.55 * size_norm + 0.20 * track_score + 0.25 * conf_norm
    chosen = int(np.argmax(score))
    return detections[chosen], centers[chosen]


# ===================================================================
# Pose extraction
# ===================================================================
def extract_poses(
    video_path: str,
    yolo_model,
    batch_size: int = 8,   # unused in ONNX path; kept for interface compatibility
    conf_thr: float = 0.05,
    device: Optional[str] = None,
) -> Tuple[np.ndarray, dict]:
    """
    Extract per-frame poses from a video using a YOLO pose ONNX model.

    Returns
    -------
    poses : ndarray, shape (T, V, 3)  --  [x, y, confidence] in pixel coords
    meta  : dict with fps, width, height, num_frames
    """
    cap = cv2.VideoCapture(str(video_path))
    if not cap.isOpened():
        raise RuntimeError(f"Cannot open video: {video_path}")

    fps = cap.get(cv2.CAP_PROP_FPS) or 30.0
    width = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH))
    height = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))

    poses_list: List[np.ndarray] = []
    prev_center: Optional[np.ndarray] = None

    while True:
        ret, frame = cap.read()
        if not ret:
            break
        detections = yolo_model.predict_frame(frame, conf_thr=conf_thr)
        kpt, prev_center = _pick_single_person(
            detections,
            prev_center=prev_center,
            conf_thr=conf_thr,
            width=width,
            height=height,
        )
        poses_list.append(kpt)

    cap.release()

    if not poses_list:
        raise RuntimeError(f"No frames in video: {video_path}")

    meta = {
        "fps": float(fps),
        "width": int(width),
        "height": int(height),
        "num_frames": len(poses_list),
    }
    return np.stack(poses_list).astype(np.float32), meta


# ===================================================================
# Preprocessing
# ===================================================================
def interpolate_poses(
    poses: np.ndarray,
    scale_factor: int = 4,
    conf_thr: float = 0.05,
    min_kpts: int = 8,
) -> np.ndarray:
    """Temporal upsampling via linear interpolation.  Input (T, V, 3)."""
    T, V, C = poses.shape
    if T <= 1 or scale_factor <= 1:
        return poses.copy()

    T_out = (T - 1) * scale_factor + 1
    out = np.zeros((T_out, V, C), dtype=np.float32)

    frame_ok = (poses[:, :, 2] >= conf_thr).sum(axis=1) >= min_kpts

    for t in range(T - 1):
        base = t * scale_factor
        out[base] = poses[t] if frame_ok[t] else 0.0

        if not (frame_ok[t] and frame_ok[t + 1]):
            continue

        a, b = poses[t], poses[t + 1]
        kp_ok = (a[:, 2] >= conf_thr) & (b[:, 2] >= conf_thr)

        for k in range(1, scale_factor):
            u = k / float(scale_factor)
            idx = base + k
            if kp_ok.any():
                out[idx, kp_ok, 0] = (1 - u) * a[kp_ok, 0] + u * b[kp_ok, 0]
                out[idx, kp_ok, 1] = (1 - u) * a[kp_ok, 1] + u * b[kp_ok, 1]
                out[idx, kp_ok, 2] = np.clip(
                    (1 - u) * a[kp_ok, 2] + u * b[kp_ok, 2], 0, 1,
                )

    out[-1] = poses[-1] if frame_ok[-1] else 0.0
    return out


def smooth_poses(
    poses: np.ndarray,
    alpha: float = 0.7,
    conf_thr: float = 0.05,
) -> np.ndarray:
    """EMA smoothing with hold-policy semantics. Input (T, V, 3)."""
    T = poses.shape[0]
    out = np.empty_like(poses)
    prev = poses[0].copy()
    out[0] = prev

    for t in range(1, T):
        kp = poses[t]
        valid = kp[:, 2] >= conf_thr
        sm = prev.copy()

        if valid.any():
            sm[valid, :2] = alpha * kp[valid, :2] + (1 - alpha) * prev[valid, :2]

        # Match CTR-GCN preprocessing: if a keypoint is missing, keep the
        # previous smoothed x/y instead of zeroing the joint out.
        sm[:, 2] = kp[:, 2]

        prev = sm
        out[t] = sm

    return out


def _uniform_sample(T_orig: int, T: int) -> np.ndarray:
    if T_orig <= 0:
        return np.zeros(T, dtype=np.int64)
    return np.linspace(0, T_orig - 1, T).astype(np.int64)


def prepare_for_model(
    poses: np.ndarray,
    meta: dict,
    fixed_t: int = 100,
) -> np.ndarray:
    """
    Convert (T, V, 3) poses to CTR-GCN input array (1, C=3, T, V, M=1) float32.
    """
    T_orig, V, C = poses.shape
    width, height = meta["width"], meta["height"]

    poses = poses.astype(np.float32, copy=True)
    poses[..., 0] /= float(width)
    poses[..., 1] /= float(height)

    if T_orig >= fixed_t:
        idx = _uniform_sample(T_orig, fixed_t)
        poses_t = poses[idx]
    else:
        poses_t = np.zeros((fixed_t, V, C), dtype=np.float32)
        poses_t[:T_orig] = poses

    # (T, V, C) -> (C, T, V) -> (C, T, V, 1) -> (1, C, T, V, 1)
    data = poses_t.transpose(2, 0, 1)[..., np.newaxis]
    return data[np.newaxis].astype(np.float32)


# ===================================================================
# Label extraction
# ===================================================================
def extract_tipper_label(filename: str) -> Tuple[Optional[str], Optional[str]]:
    """
    Extract direction and result from a tipper-style filename.

    Example: ``idapt811_sub001_DP_5_14-01-48.mp4`` -> ``('D', 'P')``
    """
    stem = Path(filename).stem
    match = re.search(r"_([DU][PFU])_", stem)
    if match:
        code = match.group(1)
        return code[0], code[1]
    return None, None


_RESULT_READABLE = {"P": "Pass", "F": "Fail", "U": "Undecided"}


# ===================================================================
# Single-video classification
# ===================================================================
def classify_video(
    video_path: str,
    yolo_model: YoloPoseONNX,
    ctr_gcn_session,
    config: ValidationConfig,
) -> Tuple[str, float]:
    """
    Run the full pipeline on one video.

    Returns ``(predicted_label, confidence)`` where label is 'Pass' or 'Fail'.
    """
    # 1. Pose extraction
    poses, meta = extract_poses(
        video_path,
        yolo_model,
        batch_size=config.yolo_batch_size,
        conf_thr=config.conf_thr,
        device=config.device,
    )

    # 2. Interpolation
    if config.do_interp:
        poses = interpolate_poses(
            poses, scale_factor=config.fps_scale, conf_thr=config.conf_thr,
        )

    # 3. Smoothing
    if config.do_smooth:
        poses = smooth_poses(poses, alpha=config.ema_alpha, conf_thr=config.conf_thr)

    # 4. Prepare input
    data = prepare_for_model(poses, meta, fixed_t=config.fixed_t)

    # 5. ONNX inference
    input_name = ctr_gcn_session.get_inputs()[0].name
    logits = ctr_gcn_session.run(None, {input_name: data})[0]  # (1, num_class)

    # 6. Softmax + argmax (pure numpy)
    shifted = logits - logits.max(axis=1, keepdims=True)
    exp = np.exp(shifted)
    probs = exp / exp.sum(axis=1, keepdims=True)
    pred = int(np.argmax(probs[0]))
    confidence = float(probs[0, pred])

    label = "Pass" if pred == 0 else "Fail"
    return label, confidence


# ===================================================================
# Batch validation
# ===================================================================
def validate_videos(
    video_entries: List[Tuple[str, str, str]],
    config: ValidationConfig,
    log: Callable[[str], None] = print,
    progress: Callable[[int, int], None] = lambda _c, _t: None,
    stop_requested: Callable[[], bool] = lambda: False,
) -> List[ValidationResult]:
    """
    Classify every renamed video.

    Parameters
    ----------
    video_entries
        List of ``(original_filename, renamed_filename, video_file_path)`` tuples.
    config
        Full validation configuration.
    log / progress / stop_requested
        Callbacks for UI integration.
    """
    (requested_yolo_backend, requested_yolo_providers), (
        requested_ctr_backend,
        requested_ctr_providers,
    ) = resolve_validation_execution(config.device, config.yolo_model_path)
    log(
        "[Validate] ONNX backend preference: "
        f"YOLO={requested_yolo_backend} ({', '.join(requested_yolo_providers)}), "
        f"CTR-GCN={requested_ctr_backend} ({', '.join(requested_ctr_providers)})"
    )
    log("[Validate] Loading YOLO model...")
    if requested_yolo_backend == "coreml_direct":
        yolo_model = YoloPoseCoreML(config.yolo_model_path)
        yolo_backend, yolo_providers = "coreml_direct", ["CoreMLDirect"]
    else:
        yolo_session, yolo_backend, yolo_providers = create_onnx_session(
            config.yolo_model_path, requested_yolo_backend, log,
        )
        yolo_model = YoloPoseONNX(config.yolo_model_path, session=yolo_session)
    log(f"[Validate] YOLO backend: {yolo_backend} ({', '.join(yolo_providers)})")

    log("[Validate] Loading CTR-GCN model...")
    ctr_gcn_session, ctr_backend, ctr_providers = create_onnx_session(
        config.ctr_gcn_checkpoint_path, requested_ctr_backend, log,
    )
    log(f"[Validate] CTR-GCN backend: {ctr_backend} ({', '.join(ctr_providers)})")

    results: List[ValidationResult] = []
    total = len(video_entries)
    log(f"[Validate] {total} videos to classify.")

    for i, (orig, renamed, path) in enumerate(video_entries):
        if stop_requested():
            log("[Validate] Cancelled by user.")
            break

        log(f"[Validate] ({i + 1}/{total}) {renamed}")
        progress(i + 1, total)

        _dir, result_code = extract_tipper_label(renamed)
        tipper_readable = _RESULT_READABLE.get(result_code, "Unknown")

        try:
            pred_label, pred_prob = classify_video(
                path, yolo_model, ctr_gcn_session, config,
            )

            if tipper_readable in ("Pass", "Fail"):
                match = tipper_readable == pred_label
            else:
                match = False

            results.append(
                ValidationResult(
                    original_video=orig,
                    renamed_video=renamed,
                    renamed_video_path=str(path),
                    tipper_label=tipper_readable,
                    predicted_label=pred_label,
                    predicted_prob=pred_prob,
                    labels_match=match,
                )
            )
        except Exception as exc:
            log(f"  [WARN] Failed: {exc}")
            results.append(
                ValidationResult(
                    original_video=orig,
                    renamed_video=renamed,
                    renamed_video_path=str(path),
                    tipper_label=tipper_readable,
                    predicted_label="Error",
                    predicted_prob=0.0,
                    labels_match=False,
                    error=str(exc),
                )
            )

    return results


# ===================================================================
# Spreadsheet output
# ===================================================================
def write_validation_report(
    results: List[ValidationResult],
    output_path: Path,
) -> None:
    """Write validation results to an Excel spreadsheet."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Validation Results"

    headers = [
        "Original Video",
        "Renamed Video",
        "Tipper Classification",
        "Model Classification",
        "Model Confidence",
        "Match",
        "Error",
    ]

    header_fill = PatternFill(fill_type="solid", fgColor="18366F")
    header_font = Font(bold=True, color="FFFFFF")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    pass_fill = PatternFill(fill_type="solid", fgColor="D4EDDA")
    fail_fill = PatternFill(fill_type="solid", fgColor="F8D7DA")
    match_fill = PatternFill(fill_type="solid", fgColor="D4EDDA")
    mismatch_fill = PatternFill(fill_type="solid", fgColor="FFF3CD")
    error_fill = PatternFill(fill_type="solid", fgColor="F8D7DA")

    for row_idx, r in enumerate(results, 2):
        ws.cell(row=row_idx, column=1, value=r.original_video)
        ws.cell(row=row_idx, column=2, value=r.renamed_video)

        tipper_cell = ws.cell(row=row_idx, column=3, value=r.tipper_label)
        if r.tipper_label == "Pass":
            tipper_cell.fill = pass_fill
        elif r.tipper_label == "Fail":
            tipper_cell.fill = fail_fill

        pred_cell = ws.cell(row=row_idx, column=4, value=r.predicted_label)
        if r.predicted_label == "Pass":
            pred_cell.fill = pass_fill
        elif r.predicted_label == "Fail":
            pred_cell.fill = fail_fill
        elif r.predicted_label == "Error":
            pred_cell.fill = error_fill

        ws.cell(row=row_idx, column=5, value=round(r.predicted_prob, 4))

        match_cell = ws.cell(
            row=row_idx, column=6, value="Yes" if r.labels_match else "No",
        )
        match_cell.fill = match_fill if r.labels_match else mismatch_fill

        if r.error:
            err_cell = ws.cell(row=row_idx, column=7, value=r.error)
            err_cell.fill = error_fill

    # Summary section
    total = len(results)
    errors = sum(1 for r in results if r.error)
    comparable = total - errors
    matches = sum(1 for r in results if r.labels_match)

    summary_row = total + 3
    ws.cell(row=summary_row, column=1, value="Summary").font = Font(bold=True)
    ws.cell(row=summary_row + 1, column=1, value=f"Total videos: {total}")
    ws.cell(
        row=summary_row + 2,
        column=1,
        value=f"Matches: {matches}/{comparable}",
    )
    if comparable > 0:
        ws.cell(
            row=summary_row + 3,
            column=1,
            value=f"Agreement rate: {matches / comparable:.1%}",
        )
    ws.cell(row=summary_row + 4, column=1, value=f"Errors: {errors}")

    # Auto-fit columns
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        max_len = max((len(str(c.value or "")) for c in col_cells), default=0)
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 45)

    ws.freeze_panes = "A2"

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
