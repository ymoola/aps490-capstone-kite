from __future__ import annotations

import csv
from dataclasses import dataclass, field
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
from typing import List, Optional, Tuple

import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from scipy.io import loadmat

from .models import TipperInfo


@dataclass
class TipperCorrection:
    date: str
    sub: str
    original_name: str
    new_name: str
    reason: str


@dataclass
class MappingEntry:
    tipper_filename: str
    date: str
    time: str
    sub: str
    shoe: str
    trial_number: Optional[float]
    angle: Optional[int]
    ice_temperature_degC: Optional[float]
    air_humidity_rh: Optional[float]
    air_temperature_degC: Optional[float]
    ultrasonic_distance_m: Optional[float]
    total_trial_time: Optional[float]
    original_video: str
    renamed_video: str
    dry_run: bool


@dataclass
class TipperMatMetadata:
    trial_number: Optional[float]
    angle: Optional[float]
    ice_temperature_degC: Optional[float]
    air_humidity_rh: Optional[float]
    air_temperature_degC: Optional[float]
    ultrasonic_distance_m: Optional[float]
    total_trial_time: Optional[float]


@dataclass
class MaaTrial:
    angle: int
    direction: str
    result: str
    trial_number: Optional[float]
    tipper_filename: str


def _safe_float(value: object) -> Optional[float]:
    try:
        return float(value)
    except Exception:
        return None


def _round_half_up_to_int(value: Optional[float]) -> Optional[int]:
    if value is None:
        return None
    try:
        return int(Decimal(str(value)).quantize(Decimal("1"), rounding=ROUND_HALF_UP))
    except Exception:
        return None


def _select_primary_mat_array(mat_data: dict) -> Optional[np.ndarray]:
    arrays = [v for v in mat_data.values() if isinstance(v, np.ndarray) and v.ndim == 2]
    if not arrays:
        return None
    return max(arrays, key=lambda a: a.size)


def _extract_mat_metadata(path: Path) -> TipperMatMetadata:
    try:
        mat_data = loadmat(path)
        mat = _select_primary_mat_array(mat_data)
        if mat is None or mat.shape[0] < 8:
            return TipperMatMetadata(None, None, None, None, None, None, None)

        trial_number = _safe_float(mat[1, 0]) if mat.shape[0] > 1 and mat.shape[1] > 0 else None
        angle = _safe_float(mat[2, 0]) if mat.shape[0] > 2 and mat.shape[1] > 0 else None
        ice_temp = _safe_float(mat[3, 0]) if mat.shape[0] > 3 and mat.shape[1] > 0 else None
        air_humidity = _safe_float(mat[4, 0]) if mat.shape[0] > 4 and mat.shape[1] > 0 else None
        air_temp = _safe_float(mat[5, 0]) if mat.shape[0] > 5 and mat.shape[1] > 0 else None
        ultrasonic = _safe_float(mat[7, 0]) if mat.shape[0] > 7 and mat.shape[1] > 0 else None

        total_trial_time: Optional[float] = None
        if mat.shape[0] > 0 and mat.shape[1] > 0:
            first_time = _safe_float(mat[0, 0])
            last_time = _safe_float(mat[0, mat.shape[1] - 1])
            if first_time is not None and last_time is not None:
                total_trial_time = last_time - first_time

        return TipperMatMetadata(
            trial_number=trial_number,
            angle=angle,
            ice_temperature_degC=ice_temp,
            air_humidity_rh=air_humidity,
            air_temperature_degC=air_temp,
            ultrasonic_distance_m=ultrasonic,
            total_trial_time=total_trial_time,
        )
    except Exception:
        return TipperMatMetadata(None, None, None, None, None, None, None)


def _shoe_and_sub_from_tipper(tipper: TipperInfo) -> Tuple[str, str]:
    parts = tipper.path.stem.split("_")
    shoe = parts[0] if parts else ""
    sub_from_name = parts[1] if len(parts) > 1 else ""
    return shoe, sub_from_name


def _time_string(time_tuple: Tuple[int, int, int]) -> str:
    hh, mm, ss = time_tuple
    if hh < 0 or mm < 0 or ss < 0:
        return ""
    return f"{hh:02}:{mm:02}:{ss:02}"


def _display_number(value: Optional[float]) -> Optional[float | int]:
    if value is None:
        return None
    try:
        if float(value).is_integer():
            return int(value)
        return float(value)
    except Exception:
        return value


def _parse_direction_result(filename: str) -> Tuple[Optional[str], Optional[str]]:
    stem = Path(filename).stem
    parts = stem.split("_")
    if len(parts) < 3:
        return None, None
    token = parts[2].strip().upper()
    if not token:
        return None, None
    direction = token[0] if token[0] in {"U", "D"} else None
    result = token[1] if len(token) >= 2 and token[1] in {"P", "F", "U"} else None
    return direction, result


def _mapping_to_maa_trial(mapping: MappingEntry) -> Optional[MaaTrial]:
    direction, result = _parse_direction_result(mapping.tipper_filename)
    if mapping.angle is None or direction not in {"U", "D"} or result not in {"P", "F", "U"}:
        return None
    return MaaTrial(
        angle=mapping.angle,
        direction=direction,
        result=result,
        trial_number=mapping.trial_number,
        tipper_filename=mapping.tipper_filename,
    )


def _compute_maa(trials: List[MaaTrial], direction: str) -> Optional[int]:
    counts_by_angle: dict[int, dict[str, int]] = {}
    for trial in trials:
        if trial.direction != direction:
            continue
        if trial.result not in {"P", "F"}:
            continue
        counts = counts_by_angle.setdefault(trial.angle, {"P": 0, "F": 0})
        counts[trial.result] += 1

    if not counts_by_angle:
        return None

    candidate_angles: List[int] = []
    for angle, counts in counts_by_angle.items():
        if counts["P"] >= 2 and counts_by_angle.get(angle + 1, {}).get("F", 0) >= 2:
            candidate_angles.append(angle)

    if candidate_angles:
        return min(candidate_angles)

    passing_angles = [angle for angle, counts in counts_by_angle.items() if counts["P"] >= 2]
    if passing_angles:
        return max(passing_angles)

    return None


def _auto_fit_columns(ws) -> None:
    for column_cells in ws.columns:
        column_letter = get_column_letter(column_cells[0].column)
        max_length = 0
        for cell in column_cells:
            if cell.value is None:
                continue
            max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 28)


def _safe_sheet_title(name: str, existing_titles: set[str]) -> str:
    invalid_chars = set('[]:*?/\\')
    sanitized = "".join("_" if ch in invalid_chars else ch for ch in name).strip()
    sanitized = sanitized or "Sheet"
    sanitized = sanitized[:31]
    candidate = sanitized
    suffix = 2
    while candidate in existing_titles:
        suffix_text = f"_{suffix}"
        candidate = f"{sanitized[: 31 - len(suffix_text)]}{suffix_text}"
        suffix += 1
    existing_titles.add(candidate)
    return candidate


def _write_maa_sheet(ws, date: str, sub: str, shoe: str, trials: List[MaaTrial]) -> None:
    title_fill = PatternFill(fill_type="solid", fgColor="DCE6F2")
    result_fill = PatternFill(fill_type="solid", fgColor="FFF5B1")
    bold = Font(bold=True)

    uphill_maa = _compute_maa(trials, "U")
    downhill_maa = _compute_maa(trials, "D")

    ws["A1"] = "MAA Summary"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A3"] = "Date"
    ws["B3"] = date
    ws["A4"] = "Participant"
    ws["B4"] = sub
    ws["A5"] = "Shoe"
    ws["B5"] = shoe
    ws["A6"] = "Uphill MAA"
    ws["B6"] = uphill_maa if uphill_maa is not None else "Insufficient data"
    ws["A7"] = "Downhill MAA"
    ws["B7"] = downhill_maa if downhill_maa is not None else "Insufficient data"
    ws["A6"].font = bold
    ws["A7"].font = bold
    ws["B6"].fill = result_fill
    ws["B7"].fill = result_fill

    summary_start = 10
    summary_headers = ["Angle", "U Pass", "U Fail", "D Pass", "D Fail"]
    for col_idx, header in enumerate(summary_headers, start=1):
        cell = ws.cell(row=summary_start, column=col_idx, value=header)
        cell.font = bold
        cell.fill = title_fill
        cell.alignment = Alignment(horizontal="center")

    angles = sorted({trial.angle for trial in trials})
    for row_idx, angle in enumerate(angles, start=summary_start + 1):
        u_pass = sum(1 for trial in trials if trial.angle == angle and trial.direction == "U" and trial.result == "P")
        u_fail = sum(1 for trial in trials if trial.angle == angle and trial.direction == "U" and trial.result == "F")
        d_pass = sum(1 for trial in trials if trial.angle == angle and trial.direction == "D" and trial.result == "P")
        d_fail = sum(1 for trial in trials if trial.angle == angle and trial.direction == "D" and trial.result == "F")
        ws.cell(row=row_idx, column=1, value=angle)
        ws.cell(row=row_idx, column=2, value=u_pass)
        ws.cell(row=row_idx, column=3, value=u_fail)
        ws.cell(row=row_idx, column=4, value=d_pass)
        ws.cell(row=row_idx, column=5, value=d_fail)

    detail_start = summary_start + len(angles) + 3
    detail_headers = ["Trial Number", "Angle", "Direction", "Result", "Tipper Filename"]
    for col_idx, header in enumerate(detail_headers, start=1):
        cell = ws.cell(row=detail_start, column=col_idx, value=header)
        cell.font = bold
        cell.fill = title_fill
        cell.alignment = Alignment(horizontal="center")

    def _trial_sort_key(trial: MaaTrial) -> tuple:
        trial_num = trial.trial_number if trial.trial_number is not None else float("inf")
        return (trial_num, trial.angle, trial.direction, trial.tipper_filename)

    for row_idx, trial in enumerate(sorted(trials, key=_trial_sort_key), start=detail_start + 1):
        ws.cell(row=row_idx, column=1, value=_display_number(trial.trial_number))
        ws.cell(row=row_idx, column=2, value=trial.angle)
        ws.cell(row=row_idx, column=3, value=trial.direction)
        ws.cell(row=row_idx, column=4, value=trial.result)
        ws.cell(row=row_idx, column=5, value=trial.tipper_filename)

    ws.freeze_panes = "A10"
    _auto_fit_columns(ws)


def _write_maa_reports(mappings: List[MappingEntry], out_dir: Path) -> None:
    grouped: dict[str, dict[str, dict[str, List[MaaTrial]]]] = {}

    for mapping in mappings:
        trial = _mapping_to_maa_trial(mapping)
        if trial is None:
            continue
        grouped.setdefault(mapping.date, {}).setdefault(mapping.sub, {}).setdefault(mapping.shoe, []).append(trial)

    if not grouped:
        return

    for date, subs in grouped.items():
        date_dir = out_dir / date
        date_dir.mkdir(parents=True, exist_ok=True)

        for sub, shoes in subs.items():
            workbook = Workbook()
            default_sheet = workbook.active
            workbook.remove(default_sheet)
            existing_titles: set[str] = set()

            for shoe, trials in sorted(shoes.items()):
                sheet_title = _safe_sheet_title(shoe, existing_titles)
                ws = workbook.create_sheet(title=sheet_title)
                _write_maa_sheet(ws, date=date, sub=sub, shoe=shoe, trials=trials)

            workbook.save(date_dir / f"{sub}.xlsx")


@dataclass
class ReportCollector:
    corrections: List[TipperCorrection] = field(default_factory=list)
    failures: List[str] = field(default_factory=list)
    mappings: List[MappingEntry] = field(default_factory=list)
    _metadata_cache: dict = field(default_factory=dict)

    def record_correction(self, date: str, sub: str, original: str, new: str, reason: str) -> None:
        self.corrections.append(TipperCorrection(date, sub, original, new, reason))

    def record_failure(self, path_label: str) -> None:
        if path_label not in self.failures:
            self.failures.append(path_label)

    def record_mapping(
        self,
        date: str,
        sub: str,
        tipper: TipperInfo,
        original_video: str,
        renamed_video: str,
        dry_run: bool,
    ) -> None:
        tipper_path = tipper.path.resolve(strict=False)
        cache_key = str(tipper_path)
        if cache_key not in self._metadata_cache:
            self._metadata_cache[cache_key] = _extract_mat_metadata(tipper_path)
        metadata: TipperMatMetadata = self._metadata_cache[cache_key]

        shoe, sub_from_name = _shoe_and_sub_from_tipper(tipper)
        effective_sub = sub or sub_from_name
        self.mappings.append(
            MappingEntry(
                tipper_filename=tipper.path.name,
                date=date,
                time=_time_string(tipper.time_tuple),
                sub=effective_sub,
                shoe=shoe,
                trial_number=metadata.trial_number,
                angle=_round_half_up_to_int(metadata.angle),
                ice_temperature_degC=metadata.ice_temperature_degC,
                air_humidity_rh=metadata.air_humidity_rh,
                air_temperature_degC=metadata.air_temperature_degC,
                ultrasonic_distance_m=metadata.ultrasonic_distance_m,
                total_trial_time=metadata.total_trial_time,
                original_video=original_video,
                renamed_video=renamed_video,
                dry_run=dry_run,
            )
        )

    def write_reports(self, out_dir: Path) -> None:
        out_dir.mkdir(parents=True, exist_ok=True)
        with (out_dir / "tipper_corrections.csv").open("w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(["date", "sub", "original_name", "new_name", "reason"])
            for c in self.corrections:
                writer.writerow([c.date, c.sub, c.original_name, c.new_name, c.reason])

        with (out_dir / "failed_folders.csv").open("w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(["folder"])
            for item in self.failures:
                writer.writerow([item])

        wb = Workbook()
        ws = wb.active
        ws.title = "Mappings"
        ws.append(
            [
                "tipper_filename",
                "date",
                "time",
                "sub",
                "shoe",
                "Trial Number",
                "Angle",
                "Ice Temperature degC",
                "Air Humidity RH",
                "Air Temperature degC",
                "Ultrasonic Distance (m)",
                "Total Trial Time",
                "original_video",
                "renamed_video",
                "dry_run",
            ]
        )
        for m in self.mappings:
            ws.append(
                [
                    m.tipper_filename,
                    m.date,
                    m.time,
                    m.sub,
                    m.shoe,
                    m.trial_number,
                    m.angle,
                    m.ice_temperature_degC,
                    m.air_humidity_rh,
                    m.air_temperature_degC,
                    m.ultrasonic_distance_m,
                    m.total_trial_time,
                    m.original_video,
                    m.renamed_video,
                    m.dry_run,
                ]
            )
        wb.save(out_dir / "video_mappings.xlsx")
        _write_maa_reports(self.mappings, out_dir / "MAA")
